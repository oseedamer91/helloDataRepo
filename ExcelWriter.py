import openpyxl
from openpyxl import Workbook
from DataIn import DataIn


# basic excel operations

def excelReadFromFile(fileName, row, col):
    book = openpyxl.load_workbook(fileName)
    sheet = book.active
    return sheet.cell(row, col)


# Read the title/ Name and their url images and total amount from excel file ,
# then return the data into object
def excelReadFile(fileName):
    book = openpyxl.load_workbook(fileName)
    sheet = book.active

    namesDataArr = []
    urlDataArr = []
    releaseDataArr = []
    totalDataArr = []

    row_count = sheet.max_row
    column_count = sheet.max_column

    r = 1
    c = 1

    for r in range(2,row_count+1):
        for c in range(1,column_count,4):
            namesDataArr.append(sheet.cell(r,c).value)
            urlDataArr.append(sheet.cell(r, c+1).value)
            releaseDataArr.append(sheet.cell(r, c+2).value)
            totalDataArr.append(sheet.cell(r, c+3).value)


    #         print(name)
    #         nameData = name
    #         urlData = url
    #         releaseData = release
    #         totalData = total
    #         namesDataArr.append(nameData)
    #         urlDataArr.append(urlData)
    #         releaseDataArr.append(releaseData)
    #         totalDataArr.append(totalData)
    # d = DataIn(namesDataArr, urlDataArr, releaseDataArr, totalDataArr)


class ExcelWriter:

    def __init__(self, excelFileName):
        self.excelName = excelFileName
        self.workbook = Workbook()
        self.sheet = self.workbook.active

    def excelWriterCell(self, column, row, value):
        self.sheet[str(column) + str(row)] = value

    def excelSave(self):
        self.workbook.save(self.excelName)


# d = {'1999': [2,3,5]}
# d2 = {'2000' : [4,4,5]}

# d2[d.keys()] = d.values()

# for k in d:
#     d2[k] = d[k]
#
# print(d2)

excelReadFile('pc.xlsx')
